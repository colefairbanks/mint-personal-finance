import pandas as pd
import numpy as np
import calendar
import openpyxl

#load csv from mint.inuit.com
df = pd.read_csv('transactions.csv')

#Converting values to negative if they are classified as "debit"
df['Amount'] = np.where(df['Transaction Type'] == 'debit', df['Amount'] * -1, df['Amount'])

#Getting rid of unnecessary columns
del df['Original Description']
del df['Transaction Type']
del df['Account Name']
del df['Labels']
del df['Notes']

#Adding a month and year column
df['Date'] =  pd.to_datetime(df['Date'], format='%m/%d/%Y')
df['Year'] = pd.DatetimeIndex(df['Date']).year
df['Month'] = pd.DatetimeIndex(df['Date']).month
df['Month'] = df['Month'].apply(lambda x: calendar.month_abbr[x])

#Arranging columns in partcular order
df = df[['Date','Year','Month','Description','Category','Amount']]

#export csv
df.to_csv('transactions_cleaned.csv',index=False)

#helper function that enables pasting dataframe into excel tab "transactions_data"
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

#executing helper function by specifying excel file name, dataframe name, sheet name, index, and where data frame should start
append_df_to_excel('PersonalFinances_template.xlsx', df, sheet_name='transactions_data', index=False, startrow=0)
